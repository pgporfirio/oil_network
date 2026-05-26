"""Generate the practitioner-spreadsheet workbooks that Chapter 5 describes.

Three xlsx files written to outputs/docs/examples/. Each workbook has two
sheets — "Before" (the analyst's starting state) and "After" (what the analyst
has to do to handle the issue) — to make visible the spreadsheet workflow that
the thesis text critiques. Values are illustrative of the order of magnitude
discussed in the thesis; the point is the workflow, not the precise figures.

Workbooks:
  ch5_2_padd5_in_transit.xlsx        — §5.2 PADD 5 in-transit anomaly
  ch5_3_hurricane_harvey.xlsx        — §5.3 Hurricane Harvey, August 2017
  ch5_4_genscape_integration.xlsx    — §5.4 Multi-source integration

Colour conventions:
  light grey header rows
  pale red    — residual / unallocated / double-count risk cells
  pale yellow — manually maintained / hand-edited cells
  pale green  — values that come from a clean canonical source
"""
from __future__ import annotations
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


def landscape_fit(ws):
    """Set the worksheet to landscape, fit-to-width, narrow margins."""
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)

OUT = Path(__file__).resolve().parent.parent / "outputs" / "docs" / "examples"

# Fills
GREY      = PatternFill("solid", fgColor="DDDDDD")
RED       = PatternFill("solid", fgColor="F4CCCC")
YELLOW    = PatternFill("solid", fgColor="FFF2CC")
GREEN     = PatternFill("solid", fgColor="D9EAD3")
BLUE      = PatternFill("solid", fgColor="CFE2F3")

# Borders
THIN  = Side(border_style="thin", color="999999")
BOX   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BOLD       = Font(bold=True)
ITALIC     = Font(italic=True, color="666666")
RED_FONT   = Font(color="9C0006", bold=True)


def cell(ws, row, col, value, *, fill=None, font=None, align=None, border=BOX, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill is not None:
        c.fill = fill
    if font is not None:
        c.font = font
    if align is not None:
        c.alignment = align
    if border is not None:
        c.border = border
    if number_format is not None:
        c.number_format = number_format
    return c


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# §5.2 — PADD 5 in-transit anomaly
# ---------------------------------------------------------------------------
def build_padd5_workbook():
    wb = Workbook()

    # Sheet 1: Before — naive PADD 5 balance, gap visible but unallocated
    ws = wb.active
    ws.title = "Before — naive balance"

    cell(ws, 1, 1, "PADD 5 monthly crude oil stock identity — representative month, July 2023",
         font=BOLD, border=None)
    cell(ws, 2, 1, "Values in thousand barrels (kbbl). EIA stock series from PSM and weekly W41.",
         font=ITALIC, border=None)

    # Header row
    cell(ws, 4, 1, "Item", fill=GREY, font=BOLD)
    cell(ws, 4, 2, "Source", fill=GREY, font=BOLD)
    cell(ws, 4, 3, "kbbl", fill=GREY, font=BOLD, align=Alignment(horizontal="right"))
    cell(ws, 4, 4, "Notes", fill=GREY, font=BOLD)

    rows = [
        ("Commercial stocks — PADD 5 (EIA total)", "EIA MCRSTP51",       55_800, "Reported published value", GREEN),
        ("",                                       "",                    None,    "",                          None),
        ("  Named tank-farm: Los Angeles",         "operator filings",    18_200, "",                          None),
        ("  Named tank-farm: San Francisco Bay",   "operator filings",    14_500, "",                          None),
        ("  Named tank-farm: Puget Sound",         "operator filings",     8_700, "",                          None),
        ("  Other named PADD-5 points",            "operator filings",     8_500, "",                          None),
        ("SPR allocated to PADD 5",                "EIA SPR table",            0, "All SPR sites are on Gulf Coast", None),
        ("Sum of named points",                    "computed",             49_900, "", GREEN),
        ("",                                       "",                    None,    "",                          None),
        ("Implied gap (EIA total − named sum)",    "computed",              5_900, "≈ 6 mmbbl — where does this go?", RED),
    ]
    for i, (item, src, val, note, fill) in enumerate(rows, start=5):
        cell(ws, i, 1, item)
        cell(ws, i, 2, src, font=ITALIC if src else None)
        cell(ws, i, 3, val, align=Alignment(horizontal="right"), number_format="#,##0;-#,##0")
        cell(ws, i, 4, note)
        if fill:
            for c in range(1, 5):
                ws.cell(row=i, column=c).fill = fill
        if item.startswith("Sum") or item.startswith("Implied"):
            for c in range(1, 5):
                ws.cell(row=i, column=c).font = BOLD
            if item.startswith("Implied"):
                ws.cell(row=i, column=3).font = RED_FONT

    cell(ws, 17, 1, "The published PADD-5 total exceeds the sum of every commercial point the analyst can name by",
         border=None, font=ITALIC)
    cell(ws, 18, 1, "about 6 million barrels every single month. No row in this sheet accounts for the gap; the",
         border=None, font=ITALIC)
    cell(ws, 19, 1, "balance does not close.", border=None, font=ITALIC)

    autosize(ws, [44, 22, 12, 50])
    landscape_fit(ws)

    # Sheet 2: After — adjustment row inserted
    ws2 = wb.create_sheet("After — adjustment row")
    cell(ws2, 1, 1, "PADD 5 stock identity with manual adjustment — same month, after analyst patch",
         font=BOLD, border=None)
    cell(ws2, 2, 1, "The 6,000 kbbl is dropped into an unallocated row so the totals reconcile.",
         font=ITALIC, border=None)

    cell(ws2, 4, 1, "Item", fill=GREY, font=BOLD)
    cell(ws2, 4, 2, "Source", fill=GREY, font=BOLD)
    cell(ws2, 4, 3, "kbbl", fill=GREY, font=BOLD, align=Alignment(horizontal="right"))
    cell(ws2, 4, 4, "Notes", fill=GREY, font=BOLD)

    rows2 = [
        ("Commercial stocks — PADD 5 (EIA total)", "EIA MCRSTP51",       55_800, "Reported published value", GREEN),
        ("",                                       "",                    None,    "",                          None),
        ("  Named tank-farm: Los Angeles",         "operator filings",    18_200, "",                          None),
        ("  Named tank-farm: San Francisco Bay",   "operator filings",    14_500, "",                          None),
        ("  Named tank-farm: Puget Sound",         "operator filings",     8_700, "",                          None),
        ("  Other named PADD-5 points",            "operator filings",     8_500, "",                          None),
        ("SPR allocated to PADD 5",                "EIA SPR table",            0, "",                          None),
        ("  Unallocated / in transit / other",     "manual plug",          5_900, "Set by analyst to close the balance — label varies between workbooks", YELLOW),
        ("Sum of decomposed points",               "computed",            55_800, "Now matches the EIA total", GREEN),
    ]
    for i, (item, src, val, note, fill) in enumerate(rows2, start=5):
        cell(ws2, i, 1, item)
        cell(ws2, i, 2, src, font=ITALIC if src else None)
        cell(ws2, i, 3, val, align=Alignment(horizontal="right"), number_format="#,##0;-#,##0")
        cell(ws2, i, 4, note)
        if fill:
            for c in range(1, 5):
                ws2.cell(row=i, column=c).fill = fill
        if item.startswith("Sum"):
            for c in range(1, 5):
                ws2.cell(row=i, column=c).font = BOLD

    cell(ws2, 16, 1, "The balance now closes, but the 5,900 kbbl sit in an opaque row that the next analyst",
         border=None, font=ITALIC)
    cell(ws2, 17, 1, "rotating onto the desk will not understand. The framework instead binds this volume to a",
         border=None, font=ITALIC)
    cell(ws2, 18, 1, "named in-transit corridor node (pipe_padd3_to_padd5), making the same value queryable",
         border=None, font=ITALIC)
    cell(ws2, 19, 1, "and labelled at the schema level.", border=None, font=ITALIC)

    autosize(ws2, [44, 22, 12, 60])
    landscape_fit(ws2)

    out = OUT / "ch5_2_padd5_in_transit.xlsx"
    wb.save(str(out))
    print(f"  {out.name}")


# ---------------------------------------------------------------------------
# §5.3 — Hurricane Harvey, August 2017
# ---------------------------------------------------------------------------
def build_harvey_workbook():
    wb = Workbook()

    # Sheet 1: Before — July 2017 baseline (normal month)
    ws = wb.active
    ws.title = "Before — July 2017 baseline"

    cell(ws, 1, 1, "PADD 3 monthly crude oil balance — July 2017 (pre-Harvey)", font=BOLD, border=None)
    cell(ws, 2, 1, "Values in thousand barrels per day (kbd) except ΔStocks in kbbl for the month.",
         font=ITALIC, border=None)

    cell(ws, 4, 1, "Item", fill=GREY, font=BOLD)
    cell(ws, 4, 2, "Sign", fill=GREY, font=BOLD)
    cell(ws, 4, 3, "Source", fill=GREY, font=BOLD)
    cell(ws, 4, 4, "Value", fill=GREY, font=BOLD, align=Alignment(horizontal="right"))

    rows = [
        ("Production",                        "+", "EIA MCRFPT3",  7_650, GREEN),
        ("Refinery runs",                     "−", "EIA WCRRIP32", 9_220, GREEN),
        ("Imports",                           "+", "EIA MCRIMP32", 3_120, GREEN),
        ("Exports",                           "−", "EIA MCREXP32", 1_010, GREEN),
        ("Movements out to PADD 1",           "−", "EIA MMRMUSP12",   95, GREEN),
        ("Movements out to PADD 2",           "−", "EIA MMRMUSP23",  475, GREEN),
        ("ΔStocks for the month (kbbl)",      "−", "EIA MCRSTP31",  +680, GREEN),
        ("",                                  "",  "",                None, None),
        ("Implied residual (kbd)",            "B", "computed",        +28, YELLOW),
    ]
    for i, (item, sign, src, val, fill) in enumerate(rows, start=5):
        cell(ws, i, 1, item)
        cell(ws, i, 2, sign)
        cell(ws, i, 3, src, font=ITALIC if src else None)
        cell(ws, i, 4, val, align=Alignment(horizontal="right"), number_format="#,##0;-#,##0")
        if fill:
            for c in range(1, 5):
                ws.cell(row=i, column=c).fill = fill
        if item.startswith("Implied"):
            for c in range(1, 5):
                ws.cell(row=i, column=c).font = BOLD

    cell(ws, 16, 1, "A normal month. The implied residual of +28 kbd is well below 1% of the throughput",
         border=None, font=ITALIC)
    cell(ws, 17, 1, "and is plausibly noise. The analyst leaves it in the adjustment cell unremarked.",
         border=None, font=ITALIC)

    autosize(ws, [42, 6, 22, 14])
    landscape_fit(ws)

    # Sheet 2: After — August 2017 with Harvey
    ws2 = wb.create_sheet("After — August 2017 (Harvey)")
    cell(ws2, 1, 1, "PADD 3 monthly crude oil balance — August 2017 (Hurricane Harvey)",
         font=BOLD, border=None)
    cell(ws2, 2, 1, "Same workbook, next month. ~25% of US refining capacity offline for parts of the month.",
         font=ITALIC, border=None)

    cell(ws2, 4, 1, "Item", fill=GREY, font=BOLD)
    cell(ws2, 4, 2, "Sign", fill=GREY, font=BOLD)
    cell(ws2, 4, 3, "Source", fill=GREY, font=BOLD)
    cell(ws2, 4, 4, "Value", fill=GREY, font=BOLD, align=Alignment(horizontal="right"))

    rows2 = [
        ("Production",                       "+", "EIA MCRFPT3",  7_605, GREEN),
        ("Refinery runs",                    "−", "EIA WCRRIP32", 7_810, GREEN),  # ~1,400 kbd drop
        ("Imports",                          "+", "EIA MCRIMP32", 2_990, GREEN),
        ("Exports",                          "−", "EIA MCREXP32",   870, GREEN),
        ("Movements out to PADD 1",          "−", "EIA MMRMUSP12",  100, GREEN),
        ("Movements out to PADD 2",          "−", "EIA MMRMUSP23",  470, GREEN),
        ("ΔStocks for the month (kbbl)",     "−", "EIA MCRSTP31", +9_360, GREEN),  # stocks built ~9.4 mmbbl
        ("",                                 "",  "",                None, None),
        ("Implied residual (kbd)",           "B", "computed",      +340, RED),
    ]
    for i, (item, sign, src, val, fill) in enumerate(rows2, start=5):
        cell(ws2, i, 1, item)
        cell(ws2, i, 2, sign)
        cell(ws2, i, 3, src, font=ITALIC if src else None)
        cell(ws2, i, 4, val, align=Alignment(horizontal="right"), number_format="#,##0;-#,##0")
        if fill:
            for c in range(1, 5):
                ws2.cell(row=i, column=c).fill = fill
        if item.startswith("Implied"):
            for c in range(1, 5):
                ws2.cell(row=i, column=c).font = BOLD
            ws2.cell(row=i, column=4).font = RED_FONT

    cell(ws2, 16, 1, "Implied residual jumps to +340 kbd — roughly an order of magnitude above the steady-state",
         border=None, font=ITALIC)
    cell(ws2, 17, 1, "median (45 kbd) reported in Table 5.4. In the spreadsheet this number falls into the same",
         border=None, font=ITALIC)
    cell(ws2, 18, 1, "unlabelled adjustment cell as the July 2017 noise. The framework instead labels it as B at",
         border=None, font=ITALIC)
    cell(ws2, 19, 1, "the PADD 3 node, traceable for any downstream consumer.", border=None, font=ITALIC)

    autosize(ws2, [42, 6, 22, 14])
    landscape_fit(ws2)

    out = OUT / "ch5_3_hurricane_harvey.xlsx"
    wb.save(str(out))
    print(f"  {out.name}")


# ---------------------------------------------------------------------------
# §5.4 — Multi-source integration (EIA + Genscape)
# ---------------------------------------------------------------------------
def build_genscape_workbook():
    wb = Workbook()

    # Sheet 1: Before — EIA only, clean monthly PADD 2 balance
    ws = wb.active
    ws.title = "Before — EIA only"

    cell(ws, 1, 1, "PADD 2 monthly crude oil balance — EIA-only workbook, representative month",
         font=BOLD, border=None)
    cell(ws, 2, 1, "Values in thousand barrels per day (kbd) except ΔStocks in kbbl for the month.",
         font=ITALIC, border=None)

    cell(ws, 4, 1, "Item", fill=GREY, font=BOLD)
    cell(ws, 4, 2, "Sign", fill=GREY, font=BOLD)
    cell(ws, 4, 3, "Source", fill=GREY, font=BOLD)
    cell(ws, 4, 4, "Value", fill=GREY, font=BOLD, align=Alignment(horizontal="right"))

    rows = [
        ("Production",                       "+", "EIA MCRFPT2",     1_530, GREEN),
        ("Refinery runs",                    "−", "EIA WCRRIP22",    3_840, GREEN),
        ("Imports — total",                  "+", "EIA MCRIMP22",    2_640, GREEN),
        ("Exports",                          "−", "EIA MCREXP22",       40, GREEN),
        ("Movements from PADD 3 to PADD 2",  "+", "EIA MMRMUSP32",     680, GREEN),
        ("Movements to PADD 1",              "−", "EIA MMRMUSP21",      45, GREEN),
        ("Movements to PADD 4",              "−", "EIA MMRMUSP24",      35, GREEN),
        ("ΔStocks for the month (kbbl)",     "−", "EIA MCRSTP21",     +850, GREEN),
        ("",                                 "",  "",                  None, None),
        ("Implied residual (kbd)",           "B", "computed",          +60, YELLOW),
    ]
    for i, (item, sign, src, val, fill) in enumerate(rows, start=5):
        cell(ws, i, 1, item)
        cell(ws, i, 2, sign)
        cell(ws, i, 3, src, font=ITALIC if src else None)
        cell(ws, i, 4, val, align=Alignment(horizontal="right"), number_format="#,##0;-#,##0")
        if fill:
            for c in range(1, 5):
                ws.cell(row=i, column=c).fill = fill
        if item.startswith("Implied"):
            for c in range(1, 5):
                ws.cell(row=i, column=c).font = BOLD

    cell(ws, 17, 1, "Clean workbook. One series per line, one cell per series, residual of +60 kbd absorbs",
         border=None, font=ITALIC)
    cell(ws, 18, 1, "noise. The analyst wants to deepen this by attributing the inter-PADD movements to",
         border=None, font=ITALIC)
    cell(ws, 19, 1, "specific pipelines using daily Genscape data. See the next sheet.",
         border=None, font=ITALIC)

    autosize(ws, [42, 6, 24, 14])
    landscape_fit(ws)

    # Sheet 2: After — EIA + Genscape, the integration burden becomes visible
    ws2 = wb.create_sheet("After — EIA + Genscape")
    cell(ws2, 1, 1, "PADD 2 monthly balance — same month, with Genscape pipeline-flow data layered on",
         font=BOLD, border=None)
    cell(ws2, 2, 1, "Genscape reports daily flow on named lines. Monthly mean shown here; each new column "
         "comes with manual subtractions to prevent double-counting.",
         font=ITALIC, border=None)

    headers = ["Item", "Sign", "Source", "Original (kbd)", "Adj 1 (kbd)", "Adj 2 (kbd)", "Net (kbd)", "Notes"]
    for j, h in enumerate(headers, start=1):
        cell(ws2, 4, j, h, fill=GREY, font=BOLD,
             align=Alignment(horizontal="center" if j > 1 else "left", wrap_text=True))

    rows2 = [
        # (item, sign, src, original, adj1, adj2, net, notes, row_fill)
        ("Production",                          "+", "EIA MCRFPT2",          1_530, None, None, 1_530, "", GREEN),
        ("Refinery runs",                       "−", "EIA WCRRIP22",         3_840, None, None, 3_840, "", GREEN),
        ("Imports — total",                     "+", "EIA MCRIMP22",         2_640, None, None, 2_640, "", GREEN),
        ("Exports",                             "−", "EIA MCREXP22",            40, None, None,    40, "", GREEN),
        ("Movements from PADD 3 to PADD 2",     "+", "EIA MMRMUSP32",          680, -250, -380,    50, "Subtract DAPL + Keystone to avoid double-count", RED),
        ("  ↳ DAPL flow (monthly mean)",        "+", "Genscape PipeFlow",       0, +250, None,   250, "Added Bakken inflow; partially overlaps with EIA", YELLOW),
        ("  ↳ Keystone flow",                   "+", "Genscape PipeFlow",       0, +380, None,   380, "Added Canadian inflow; overlaps with EIA imports — re-check", YELLOW),
        ("  ↳ Spearhead flow (out to Cushing)", "−", "Genscape PipeFlow",       0,  +90, None,    90, "Cushing-outbound; sign convention manual", YELLOW),
        ("  ↳ Residual / un-named pipelines",   "+", "manual plug",             0,  +60, None,    60, "≈9% of capacity Genscape doesn't cover; recomputed monthly", YELLOW),
        ("Movements to PADD 1",                 "−", "EIA MMRMUSP21",           45, None, None,    45, "", GREEN),
        ("Movements to PADD 4",                 "−", "EIA MMRMUSP24",           35, None, None,    35, "", GREEN),
        ("ΔStocks for the month (kbbl)",        "−", "EIA MCRSTP21",          +850, None, None, +850, "", GREEN),
        ("",                                    "",  "",                      None, None, None,  None, "", None),
        ("Implied residual (kbd) — original",   "B", "computed before Genscape",  60, None, None,    60, "", YELLOW),
        ("Implied residual (kbd) — with Genscape", "B", "computed after Genscape", None, None, None, -45, "Should be ≈60 if integration is clean; gap signals a reconciliation problem", RED),
        ("Σ Genscape named pipelines vs EIA inter-PADD", "check", "computed",  None, None, None, 720 - 680, "Genscape sum exceeds EIA aggregate — analyst must investigate", RED),
    ]
    for i, row in enumerate(rows2, start=5):
        item, sign, src, orig, adj1, adj2, net, notes, fill = row
        cell(ws2, i, 1, item)
        cell(ws2, i, 2, sign)
        cell(ws2, i, 3, src, font=ITALIC if src else None)
        cell(ws2, i, 4, orig, align=Alignment(horizontal="right"), number_format="#,##0;-#,##0")
        cell(ws2, i, 5, adj1, align=Alignment(horizontal="right"), number_format="+#,##0;-#,##0")
        cell(ws2, i, 6, adj2, align=Alignment(horizontal="right"), number_format="+#,##0;-#,##0")
        cell(ws2, i, 7, net, align=Alignment(horizontal="right"), number_format="#,##0;-#,##0")
        cell(ws2, i, 8, notes, align=Alignment(wrap_text=True, vertical="top"))
        if fill:
            for c in range(1, 9):
                ws2.cell(row=i, column=c).fill = fill
        if item.startswith("Implied") or item.startswith("Σ"):
            for c in range(1, 9):
                ws2.cell(row=i, column=c).font = BOLD
            if "Genscape" in str(notes) or "investigate" in str(notes) or "reconciliation" in str(notes):
                ws2.cell(row=i, column=7).font = RED_FONT

    # Closing commentary
    rstart = 22
    notes = [
        "Every Genscape line costs the analyst three things: a new column, a manual subtraction from the",
        "EIA inter-PADD column to prevent double-counting, and a sign-convention decision. The residual-",
        "pipelines column for the 9% Genscape doesn't cover has to be recomputed each month. The cross-",
        "check at the bottom (ΣGenscape vs EIA aggregate) is the only protection against silent drift, and",
        "it has to be hand-coded. In the framework, Genscape values TS-bind to the named pipeline node and",
        "EIA aggregates bind to the PADD-view node; v_aggregation_consistency does the cross-check by ",
        "construction.",
    ]
    for k, line in enumerate(notes):
        cell(ws2, rstart + k, 1, line, border=None, font=ITALIC)

    autosize(ws2, [44, 6, 24, 14, 12, 12, 12, 50])
    landscape_fit(ws2)

    out = OUT / "ch5_4_genscape_integration.xlsx"
    wb.save(str(out))
    print(f"  {out.name}")


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    print(f"Writing example workbooks to {OUT}:")
    build_padd5_workbook()
    build_harvey_workbook()
    build_genscape_workbook()


if __name__ == "__main__":
    main()
